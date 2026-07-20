"""Excel export for extracted invoices."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font

from docai.models.schemas import Invoice

HEADER_FONT = Font(bold=True)


def invoices_to_xlsx(invoices: list[Invoice]) -> bytes:
    wb = Workbook()

    # Sheet 1: summary
    ws = wb.active
    ws.title = "Документы"
    headers = ["Тип", "Номер", "Дата", "Поставщик", "БИН поставщика",
               "Покупатель", "БИН покупателя", "Подытог", "НДС", "Итого", "Валюта"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for inv in invoices:
        ws.append([
            inv.doc_type, inv.doc_number, inv.doc_date,
            inv.supplier_name, inv.supplier_bin,
            inv.buyer_name, inv.buyer_bin,
            float(inv.subtotal) if inv.subtotal is not None else None,
            float(inv.vat_amount) if inv.vat_amount is not None else None,
            float(inv.total), inv.currency,
        ])

    # Sheet 2: line items
    ws2 = wb.create_sheet("Позиции")
    ws2.append(["Документ", "Наименование", "Кол-во", "Ед.", "Цена", "Сумма"])
    for cell in ws2[1]:
        cell.font = HEADER_FONT
    for inv in invoices:
        doc_ref = inv.doc_number or inv.doc_date or inv.supplier_name
        for li in inv.line_items:
            ws2.append([
                doc_ref, li.name,
                float(li.quantity) if li.quantity is not None else None,
                li.unit,
                float(li.price) if li.price is not None else None,
                float(li.total),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
