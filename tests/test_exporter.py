from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from docai.core.exporter import invoices_to_xlsx
from docai.models.schemas import Invoice, LineItem


def test_export_produces_valid_xlsx():
    inv = Invoice(
        doc_type="счёт-фактура",
        doc_number="42",
        supplier_name="ТОО Тест",
        line_items=[LineItem(name="Товар", quantity=Decimal(2), total=Decimal("500"))],
        total=Decimal("500"),
    )
    data = invoices_to_xlsx([inv])
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Документы", "Позиции"]
    ws = wb["Документы"]
    assert ws.cell(row=2, column=2).value == "42"
    assert wb["Позиции"].cell(row=2, column=2).value == "Товар"
